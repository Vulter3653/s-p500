"""Load and validate the official Brysbaert et al. (2014) XLSX supplement."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DICTIONARY_ROOT = ROOT / "references/dictionaries/brysbaert_concreteness"
DEFAULT_PATH = DICTIONARY_ROOT / "original_source_files/concreteness_ratings_brysbaert_et_al_2014.xlsx"
ANALYSIS_PATH = DICTIONARY_ROOT / "analysis_ready_dictionary/brysbaert_concreteness_analysis_ready.csv"
EXPECTED_SHA256 = "1673ead761e28833a40e82c0d20f10782955ced9366d600eafeefee0f2254545"
OFFICIAL_PAGE = "https://link.springer.com/article/10.3758/s13428-013-0403-5"
DIRECT_URL = "https://media.springernature.com/original/springer-static/esm/art%3A10.3758%2Fs13428-013-0403-5/MediaObjects/13428_2013_403_MOESM1_ESM.xlsx"
RELEASE = "Brysbaert-Warriner-Kuperman-2014"
LOADER_VERSION = "1.0.0"
REQUIRED = {"Word", "Bigram", "Conc.M", "Conc.SD", "Unknown", "Total", "Percent_known", "SUBTLEX"}


class DictionaryValidationError(ValueError):
    pass


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_entry(value: str) -> str:
    return " ".join(unicodedata.normalize("NFKC", str(value)).strip().lower().split())


def load_dictionary(path: Path = DEFAULT_PATH, expected_sha256: str = EXPECTED_SHA256,
                    expected_count: int = 39954, expected_single: int = 37058,
                    expected_bigram: int = 2896,
                    write_analysis_file: bool = False) -> tuple[dict, dict]:
    if not path.is_file():
        raise FileNotFoundError(f"official Brysbaert dictionary missing: {path}")
    actual_sha = sha256_file(path)
    if expected_sha256 and actual_sha != expected_sha256:
        raise DictionaryValidationError(
            f"Brysbaert SHA-256 mismatch: expected {expected_sha256}, got {actual_sha}"
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    missing = REQUIRED - set(headers)
    if missing:
        raise DictionaryValidationError(f"missing required columns: {sorted(missing)}")
    positions = {name: headers.index(name) for name in headers}
    entries = {}
    single_count = bigram_count = empty_count = 0
    score_min = math.inf
    score_max = -math.inf
    na_counts = {name: 0 for name in headers}
    for source_row_number, values in enumerate(rows, start=2):
        raw_word = values[positions["Word"]]
        word = normalize_entry(raw_word) if raw_word is not None else ""
        if not word:
            empty_count += 1
            continue
        if word in entries:
            raise DictionaryValidationError(f"duplicate normalized entry: {word}")
        try:
            score = float(values[positions["Conc.M"]])
            standard_deviation = float(values[positions["Conc.SD"]])
        except (TypeError, ValueError) as error:
            raise DictionaryValidationError(f"non-numeric score at row {source_row_number}") from error
        if not 1 <= score <= 5:
            raise DictionaryValidationError(f"score outside 1-5 at row {source_row_number}")
        bigram = int(values[positions["Bigram"]])
        single_count += int(bigram == 0)
        bigram_count += int(bigram == 1)
        score_min = min(score_min, score)
        score_max = max(score_max, score)
        for name, position in positions.items():
            na_counts[name] += int(values[position] is None)
        entries[word] = {
            "dictionary_entry": str(raw_word), "normalized_entry": word,
            "entry_type": "two_word_expression" if bigram else "single_word",
            "score": score, "standard_deviation": standard_deviation,
            "unknown": values[positions["Unknown"]], "total": values[positions["Total"]],
            "percent_known": values[positions["Percent_known"]],
            "subtlex": values[positions["SUBTLEX"]],
            "dictionary_row_number": source_row_number,
        }
    if empty_count:
        raise DictionaryValidationError(f"empty dictionary entries: {empty_count}")
    if len(entries) != expected_count:
        raise DictionaryValidationError(
            f"entry count mismatch: expected {expected_count}, got {len(entries)}"
        )
    if (single_count, bigram_count) != (expected_single, expected_bigram):
        raise DictionaryValidationError(f"single/bigram mismatch: {single_count}/{bigram_count}")
    metadata = {
        "dictionary_name": "Concreteness ratings for 40 thousand generally known English word lemmas",
        "dictionary_release": RELEASE, "doi": "10.3758/s13428-013-0403-5",
        "official_page_url": OFFICIAL_PAGE, "direct_download_url": DIRECT_URL,
        "local_path": str(path.relative_to(ROOT)) if path.is_relative_to(ROOT) else str(path),
        "sha256": actual_sha, "row_count": len(entries), "column_count": len(headers),
        "columns": headers, "single_word_count": single_count,
        "two_word_expression_count": bigram_count, "score_min": score_min,
        "score_max": score_max, "na_counts": na_counts,
        "redistribution_status": "research_use_permitted_redistribution_unclear",
        "loader_version": LOADER_VERSION,
    }
    if write_analysis_file:
        ANALYSIS_PATH.parent.mkdir(parents=True, exist_ok=True)
        fields = [
            "dictionary_entry", "normalized_entry", "entry_type", "concreteness_score",
            "concreteness_standard_deviation", "unknown_response_count",
            "total_response_count", "percent_known", "subtlex_frequency",
            "dictionary_row_number",
        ]
        with ANALYSIS_PATH.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
            writer.writeheader()
            for entry in entries.values():
                writer.writerow({
                    "dictionary_entry": entry["dictionary_entry"],
                    "normalized_entry": entry["normalized_entry"],
                    "entry_type": entry["entry_type"],
                    "concreteness_score": entry["score"],
                    "concreteness_standard_deviation": entry["standard_deviation"],
                    "unknown_response_count": entry["unknown"],
                    "total_response_count": entry["total"],
                    "percent_known": entry["percent_known"],
                    "subtlex_frequency": entry["subtlex"],
                    "dictionary_row_number": entry["dictionary_row_number"],
                })
    return entries, metadata


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--validate-only", action="store_true")
    parser.add_argument("--write-analysis-file", action="store_true")
    arguments = parser.parse_args()
    _, info = load_dictionary(write_analysis_file=arguments.write_analysis_file)
    print(json.dumps(info, ensure_ascii=False, sort_keys=True))
